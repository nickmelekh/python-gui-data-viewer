import sys
import os
import pandas as pd
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtCore import QCoreApplication, Qt
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt
from openpyxl import load_workbook

class Cols(object):
    def __init__(self):
        super().__init__()
        self.cols = []

class Data(object):
    def __init__(self, fileName, data):
        super().__init__()
        self.fileName = fileName
        self.data = data

class HeadersWindow(QWidget):
    def __init__(self, fileName, data):
        super().__init__()
        self.headers = [['Время', 'I нагр. max(10s), А', 'U ввод1 (10s), В', 'P1 сред(10s), кВт*ч', 'Total P1 потр. , кВт*ч', 'Q1 потр.сред. (10s), кВар*ч', 'Q1 генерир. Сред. (10s), кВар*ч', 'Total Q1, кВар*ч', 'S1 сред. (10s) , ВА', 'PF  сред. 10s (cos φ)', 'F 1 среднее 10s, Hz', 'I нагр. max(10s), А2', 'U ввод2(10s), В', 'P2сред(10s), кВт*ч', 'Total P2потр. , кВт*ч', 'Q2потр.сред. (10s), кВар*ч', 'Q2генерир. Сред. (10s), кВар*ч', 'Total Q1, кВар*ч3', 'S2сред. (10s) , ВА', 'PF  сред. 10s (cos φ)4', 'F 1 среднее 10s, Hz5', 'U Вых, В', 'V м/с сред 10 сек', 'Выходной ток', 'Угол поворота', 'Направление ветра', 'U вых, В6', 'I вых, А', 'P in, Вт средняя 10 s'], ['Час', 'Средне – часовое значение тока нагрузки − ∑ показ./ 360', 'Максимальное 10-ти секундное значение тока нагрузки', 'Средне – часовое значение напряжение на вводе − ∑ показ./ 360', 'Максимальное 10-ти секундное значение напряжения', 'Минимальное 10-ти секундное значение напряжения', 'Средне – часовое значение потребляемой активной мощности - ∑ показ./ 360', 'Максимальное 10-ти секундное значение потребляемой активной мощности', 'Средне – часовое значение потребляемой реактивной мощности − ∑ показ./ 360', 'Максимальное 10-ти секундное значение потребляемой реактивной мощности', 'Средне – часовое значение потребляемой полной мощности − ∑ показ./ 360', 'Максимальное 10-ти секундное значение потребляемой полной мощности']]
        self.filename = fileName
        self.data = data
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.filename + ' headers')
        grid = QGridLayout()
        grid.setSpacing(5)
        titleCombo = QLabel('Choose default headers:')

        self.headersCombo = QComboBox(self)
        itemsList = []
        for i in range(len(self.headers)):
            itemsList.append('; '.join(self.headers[i])[:70] + '...')
        self.headersCombo.addItems(itemsList)

        titleEdit = QLabel('Write new headers, divided by ; :')

        self.headersEdit = QLineEdit(self)

        chbtn = QPushButton('Choose', self)
        chbtn.clicked.connect(self.set_headers)

        addbtn = QPushButton('Add', self)
        addbtn.clicked.connect(self.apply_headers)

        savebtn = QPushButton('Add and save', self)
        savebtn.setEnabled(False)
        # qbtn.clicked.connect(QCoreApplication.instance().quit)

        cancelbtn = QPushButton('Cancel', self)
        cancelbtn.clicked.connect(self.quit)

        # headersCombo.activated[str].connect(self.onActivatedheaders)

        grid.addWidget(titleCombo, 0, 0)
        grid.addWidget(self.headersCombo, 1, 0, 1, 3)
        grid.addWidget(chbtn, 2, 2)
        grid.addWidget(titleEdit, 3, 0)
        grid.addWidget(self.headersEdit, 4, 0, 4, 3)
        grid.addWidget(addbtn, 8, 0)
        grid.addWidget(savebtn, 8, 1)
        grid.addWidget(cancelbtn, 8, 2)

        self.setLayout(grid)
        self.show()

    def quit(self):
        self.close()

    def set_headers(self):
        self.headersEdit.setText('; '.join(self.headers[self.headersCombo.currentIndex()]))

    def apply_headers(self):
        headers = self.headersEdit.text()
        cols = headers.split(';')
        print(len(cols), len(self.data.columns))
        if len(cols) == len(self.data.columns):
            newCols.cols = cols
            window.tables[self.filename].apply_headers()
            self.close()
        else:
            pass

class GraphWindow(QWidget):
    def __init__(self, fileName, data):
        super().__init__()
        self.setWindowTitle(fileName + ' graph')
        self.filename = fileName
        self.data = data
        self.initUI()

    def initUI(self):
        sizeObject = QDesktopWidget().screenGeometry(-1)
        self.setFixedSize(sizeObject.width(), sizeObject.height() - 100)
        # self.resize(1200, 800)

        title = QLabel('Title')
        xt = QLabel('X ticks')
        yt = QLabel('Y ticks')
        startValue = QLabel('Start value')
        endValue = QLabel('End value')

        titleEdit = QLineEdit()
        startEdit = QLineEdit()
        endedit = QLineEdit()

        ytCombo = QComboBox(self)
        itemsList = []
        for i in range(len(self.data.columns)):
            itemsList.append(str(self.data.columns[i]))
        ytCombo.addItems(itemsList)
        ytCombo.setCurrentIndex(2)
        xtCombo = QComboBox(self)
        xtCombo.addItems(itemsList)

        xtCombo.activated[str].connect(self.onActivatedxt)
        ytCombo.activated[str].connect(self.onActivatedyt)

        # qbtn = QPushButton('Save', self)
        # # qbtn.clicked.connect(QCoreApplication.instance().quit)
        # qbtn.resize(qbtn.sizeHint())
        # qbtn.setToolTip('Save plot as .PNG')

        self.m = PlotCanvas(self, self.data)

        grid = QGridLayout()
        grid.setSpacing(5)

        grid.addWidget(title, 0, 0)
        grid.addWidget(titleEdit, 0, 1)
        grid.addWidget(xt, 1, 0)
        grid.addWidget(xtCombo, 1, 1)
        grid.addWidget(yt, 2, 0)
        grid.addWidget(ytCombo, 2, 1)
        # grid.addWidget(startValue, 4, 0)
        # grid.addWidget(startEdit, 4, 1)
        # grid.addWidget(endValue, 5, 0)
        # grid.addWidget(endedit, 5, 1)
        # grid.addWidget(qbtn, 15, 0)

        grid.addWidget(self.m, 0, 2, 16, 16)
        grid.setColumnStretch(2, 16)
        self.setLayout(grid)
        self.show()

    def onActivatedxt(self, text):
        self.m.x = text
        self.m.plot()

    def onActivatedyt(self, text):
        self.m.y = text
        self.m.plot()

class PlotCanvas(FigureCanvas):
    def __init__(self, filename, data, parent = None, width = 4, height = 4, dpi = 100):
        fig = Figure(figsize = (width, height), dpi = dpi)
        self.axes = fig.add_subplot(111)
        self.data = data
        self.x = self.data.columns[0]
        self.y = self.data.columns[2]
        FigureCanvas.__init__(self, fig)
        self.setParent(parent)
        FigureCanvas.setSizePolicy(self, QSizePolicy.Expanding, QSizePolicy.Expanding)
        FigureCanvas.updateGeometry(self)
        self.plot()

    def plot(self):
        ax = self.figure.add_subplot(111)
        ax.cla()
        x = self.data[self.x].values[0:50]
        y = self.data[self.y].values[0:50]
        ax.set_autoscale_on(True)
        ax.plot(x, y)
        ax.tick_params(axis = 'x', rotation = 90, labelsize = 6)
        self.draw()

class TableWindow(QMainWindow):
    def __init__(self, fileName, data):
        super(TableWindow, self).__init__()
        self.resize(800, 600)
        self.setWindowTitle(fileName)
        self.filename = fileName
        self.data = data
        self.data.columns = [str(x) for x in range(data.shape[1])]
        self.level = 'raw'
        self.graphs = []
        self.reports = []
        self.headers = {}
        # self.path = str(self.filename[:-4] + '.xlsx')
        # self.writer = pd.ExcelWriter(self.path, engine = 'xlsxwriter')
        self.initUI()

    def initUI(self):
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        grid_layout = QGridLayout()
        central_widget.setLayout(grid_layout)

        tb = self.addToolBar('Menu')
        tb.setToolButtonStyle(3)
        tb.setMovable(False)

        createGraph = QAction(QIcon('graph.png'), 'New graph', self)
        createGraph.triggered.connect(self.create_graph)
        tb.addAction(createGraph)

        addHeaders = QAction(QIcon('header.png'), 'Add headers', self)
        addHeaders.triggered.connect(self.add_all_headers)
        tb.addAction(addHeaders)

        addReport = QAction(QIcon('report.png'), 'Report', self)
        addReport.triggered.connect(self.calculate_data1)
        tb.addAction(addReport)

        saveTable = QAction(QIcon('save.png'), 'Save', self)
        saveTable.triggered.connect(self.save_data)
        tb.addAction(saveTable)

        closeTable = QAction(QIcon('close.png'), 'Close', self)
        closeTable.triggered.connect(self.quit)
        tb.addAction(closeTable)

        self.table = QTableWidget()
        convert_pandas_to_widget(self.table, self.data)
        self.table.setHorizontalHeaderLabels(str(i) for i in range(self.data.shape[1]))
        # self.table.horizontalHeader().sectionDoubleClicked.connect(self.change_horizontal_header)

        grid_layout.addWidget(self.table)

    def save_data(self, sheet):
        if self.level == 'raw':
            self.data.to_excel(self.filename[:-4] + '.xlsx', sheet_name = self.level)
        else:
            book = load_workbook(self.filename[:-4] + '.xlsx')
            # book.create_sheet(self.level)
            writer = pd.ExcelWriter(self.filename[:-4] + '.xlsx', engine = 'openpyxl')
            writer.book = book
            # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            self.data.to_excel(writer, self.level)
            writer.save()
            # wb = load_workbook(self.filename[:-4] + '.xlsx')
            # wb.create_sheet(self.level)
            # writer = pd.ExcelWriter(self.filename[:-4] + '.xlsx')
            # self.data.to_excel(writer, self.level, index = False)
            # wb.save(self.filename[:-4] + '.xlsx')
            # writer = pd.ExcelWriter(self.filename[:-4] + '.xlsx', engine='openpyxl')
            # workbook  = writer.book
            # worksheet = workbook.create_sheet(self.level)
            # self.data.to_excel(writer, sheet_name = self.level)
            # writer.save()

    def change_horizontal_header(self, index):
        oldHeader = self.table.horizontalHeaderItem(index).text()
        newHeader, ok = QInputDialog.getText(self,
                                                      'Change header label for column %d' % index,
                                                      'Header:',
                                                       QLineEdit.Normal,
                                                       oldHeader)
        if ok:
            self.table.horizontalHeaderItem(index).setText(newHeader)

    def add_all_headers(self):
        # dlg = QInputDialog(self)
        # dlg.setInputMode(QInputDialog.TextInput)
        # dlg.setLabelText('Write headers splitted by ; :')
        # dlg.resize(500, 58)
        # ok = dlg.exec_()
        # headers = dlg.textValue()
        # cols = headers.split(';')
        # if len(cols) == len(self.data.columns):
        #     self.data.columns = cols
        #     for i in range(len(self.data.columns)):
        #         self.table.horizontalHeaderItem(i).setText(cols[i])
        #         self.table.resizeColumnsToContents()
        # else:
        #     print('try again')
        headersWindow = HeadersWindow(self.filename, self.data)
        self.headers[self.filename] = headersWindow
        self.headers[self.filename].show()

    def apply_headers(self):
        global newCols
        self.data.columns = newCols.cols
        for i in range(len(self.data.columns)):
            self.table.horizontalHeaderItem(i).setText(newCols.cols[i])
            self.table.resizeColumnsToContents()

    def create_graph(self):
        graphWindow = GraphWindow(self.filename, self.data)
        self.graphs.append(graphWindow)
        self.graphs[-1].show()

    def calculate_data1(self):
        if self.data.shape[1] > 11:
            start_time = self.data[self.data.columns[0]].values[0][4]
            hour_measures = {}
            h = 0
            j = 0
            for i in range(1, len(self.data[self.data.columns[0]].values)):
                if self.data[self.data.columns[0]].values[i][4] != start_time:
                    h += 1
                    start_time = self.data[self.data.columns[0]].values[i][4]
                    hour_measures[h] = {0: h,
                                        1: round(self.data[self.data.columns[1]][j:i].sum() / 360, 3),
                                        2: self.data[self.data.columns[1]][j:i].max(),
                                        3: round(self.data[self.data.columns[1]][j:i].sum() / 360, 3),
                                        4: self.data[self.data.columns[2]][j:i].max(),
                                        5: self.data[self.data.columns[2]][j:i].min(),
                                        6: round(self.data[self.data.columns[3]][j:i].sum() / 360, 3),
                                        7: self.data[self.data.columns[7]][j:i].max(),
                                        8: round(self.data[self.data.columns[5]][j:i].sum() / 360, 3),
                                        9: self.data[self.data.columns[7]][j:i].max(),
                                        10: round(self.data[self.data.columns[10]][j:i].sum() / 360, 3),
                                        11: self.data[self.data.columns[10]][j:i].max(),
                                       }
                    j = i

            self.data = pd.DataFrame.from_dict(hour_measures, orient = 'index')
            indexes = [x for x in range(self.data.shape[0])]
            self.data = self.data.reindex(indexes)
            self.data = self.data.shift(periods = -1)
            self.data.columns = [str(x) for x in range(self.data.shape[1])]
            self.update_table(self.data)
            self.level = 'report1'

    def update_table(self, data):
        convert_pandas_to_widget(self.table, self.data)
        self.table.setHorizontalHeaderLabels(str(i) for i in range(self.data.shape[1]))

    def quit(self):
        # self.writer.close()
        self.close()

class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.tables = {}
        self.initUI()

    def initUI(self):
        self.center()
        self.setWindowTitle('Electra')
        self.setFixedSize(200, 80)

        tb = self.addToolBar('Menu')
        tb.setToolButtonStyle(3)
        tb.setMovable(False)
        tb.setFixedWidth(240)
        tb.setFixedHeight(80)

        openFile = QAction(QIcon('file.png'), 'Open', self)
        openFile.triggered.connect(self.open_file_dialog)
        tb.addAction(openFile)

        openSettings = QAction(QIcon('settings.png'), 'Settings', self)
        # openFile.triggered.connect(self.open_file_dialog)
        tb.addAction(openSettings)

        openHelp = QAction(QIcon('help.png'), 'Help', self)
        # openFile.triggered.connect(self.open_file_dialog)
        tb.addAction(openHelp)

        exitAct = QAction(QIcon('close.png'), 'Exit', self)
        exitAct.setShortcut('Ctrl+Q')
        exitAct.triggered.connect(qApp.quit)
        tb.addAction(exitAct)

        self.show()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def open_file_dialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","All Files (*);;Python Files (*.py)", options=options)
        if fileName:
            data = Data(fileName, data = read_Csvfile(fileName))
            tableWindow = TableWindow(fileName, data.data)
            self.tables[fileName] = tableWindow
            self.tables[fileName].show()

def read_Csvfile(fileName):
    data = pd.read_csv(fileName, header = None)
    data.drop(data.columns[len(data.columns)-1], axis=1, inplace=True)
    return data

def read_Csvfile_with_header(fileName, headers):
    # print(headers)
    data = pd.read_csv(fileName, header = None, names = headers)
    return data

def convert_pandas_to_widget(widget, data):
    headers = data.columns.values.tolist()
    table = widget
    table.clear()
    table.setColumnCount(len(headers))
    for i, row in data.iterrows():
        table.setRowCount(table.rowCount() + 1)
        for j in range(table.columnCount()):
            table.setItem(i, j, QTableWidgetItem(str(row[j])))
    table.resizeColumnsToContents()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    path = os.path.join(os.path.dirname(sys.modules[__name__].__file__), 'icon_2.png')
    app.setWindowIcon(QIcon(path))
    newCols = Cols()
    window = App()
    sys.exit(app.exec_())
